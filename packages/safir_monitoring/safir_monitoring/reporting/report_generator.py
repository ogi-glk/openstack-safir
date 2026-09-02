"""Haftalik rapor verisi olusturma modulu.

MySQL'deki host_metrics_hourly ve vm_metrics_hourly tablolarindan
veri cekip rapor icin gerekli aggregate ve grafikleri olusturur.
"""

import base64
from io import BytesIO
from datetime import datetime, timedelta

import pandas as pd
import matplotlib
matplotlib.use('Agg')  # GUI olmadan calistir
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from oslo_log import log as logging

LOG = logging.getLogger(__name__)


async def get_host_metrics_summary(db, days=7):
    """Host bazli CPU/Memory/Disk avg/max/min ozetini MySQL'den ceker."""
    cutoff = datetime.now() - timedelta(days=days)

    query = """
        SELECT nodename,
               ROUND(AVG(cpu_percent), 1) as cpu_avg,
               ROUND(MAX(cpu_percent), 1) as cpu_max,
               ROUND(MIN(cpu_percent), 1) as cpu_min,
               ROUND(AVG(memory_percent), 1) as memory_avg,
               ROUND(MAX(memory_percent), 1) as memory_max,
               ROUND(MIN(memory_percent), 1) as memory_min,
               ROUND(AVG(disk_percent), 1) as disk_avg,
               ROUND(MAX(disk_percent), 1) as disk_max,
               ROUND(MIN(disk_percent), 1) as disk_min
        FROM host_metrics_hourly
        WHERE timestamp >= :cutoff AND nodename IS NOT NULL
        GROUP BY nodename
        ORDER BY nodename
    """
    rows = await db.fetch_all(query=query, values={"cutoff": cutoff})
    return [dict(r._mapping) for r in rows]


async def get_placement_summary(db, days=7):
    """Placement verisi ozetini MySQL'den ceker."""
    cutoff = datetime.now() - timedelta(days=days)

    query = """
        SELECT nodename,
               ROUND(AVG(placement_vcpu_used), 0) as vcpu_used,
               ROUND(AVG(placement_vcpu_total), 0) as vcpu_total,
               ROUND(AVG(placement_memory_used_mb / 1024), 1) as memory_used_gb,
               ROUND(AVG(placement_memory_total_mb / 1024), 1) as memory_total_gb,
               ROUND(AVG(placement_disk_used_gb), 0) as disk_used_gb,
               ROUND(AVG(placement_disk_total_gb), 0) as disk_total_gb
        FROM host_metrics_hourly
        WHERE timestamp >= :cutoff AND placement_vcpu_total IS NOT NULL
        GROUP BY nodename
        ORDER BY nodename
    """
    rows = await db.fetch_all(query=query, values={"cutoff": cutoff})
    return [dict(r._mapping) for r in rows]


async def get_rightsizing_summary(db, days=7):
    """VM rightsizing ozetini MySQL'den ceker."""
    cutoff = datetime.now() - timedelta(days=days)

    query = """
        SELECT instance_id, instance_name, project_id,
               ROUND(AVG(avg_cpu), 2) as avg_cpu,
               ROUND(MAX(max_cpu), 2) as max_cpu,
               ROUND(AVG(avg_memory), 2) as avg_memory,
               ROUND(MAX(max_memory), 2) as max_memory,
               MAX(allocated_vcpu) as allocated_vcpu,
               ROUND(MAX(allocated_memory_bytes) / (1024*1024*1024), 1) as allocated_memory_gb
        FROM vm_metrics_hourly
        WHERE timestamp >= :cutoff
        GROUP BY instance_id, instance_name, project_id
        ORDER BY instance_name
    """
    rows = await db.fetch_all(query=query, values={"cutoff": cutoff})

    idle = []
    over = []
    under = []
    right_sized = 0

    for r in rows:
        vm = dict(r._mapping)
        avg_cpu = vm.get("avg_cpu") or 0
        avg_mem = vm.get("avg_memory") or 0
        max_cpu = vm.get("max_cpu") or 0
        max_mem = vm.get("max_memory") or 0

        if avg_cpu < 5 and avg_mem < 10:
            idle.append(vm)
        elif avg_cpu > 80 or avg_mem > 80:
            under.append(vm)
        elif max_cpu < 30 and max_mem < 30:
            over.append(vm)
        else:
            right_sized += 1

    return {
        "total": len(rows),
        "idle": idle,
        "over_provisioned": over,
        "under_provisioned": under,
        "right_sized_count": right_sized,
    }


async def get_weekly_trend_data(db, days=7):
    """Gunluk trend verisi — grafik icin."""
    cutoff = datetime.now() - timedelta(days=days)

    query = """
        SELECT DATE(timestamp) as day,
               ROUND(AVG(cpu_percent), 2) as cpu_avg,
               ROUND(AVG(memory_percent), 2) as memory_avg,
               ROUND(AVG(disk_percent), 2) as disk_avg
        FROM host_metrics_hourly
        WHERE timestamp >= :cutoff AND cpu_percent IS NOT NULL
        GROUP BY DATE(timestamp)
        ORDER BY day
    """
    rows = await db.fetch_all(query=query, values={"cutoff": cutoff})
    return [dict(r._mapping) for r in rows]


def generate_trend_charts(trend_data):
    """Haftalik trend grafiklerini base64 olarak olusturur."""
    if not trend_data:
        return {}

    dates = [r["day"] for r in trend_data]
    charts = {}

    metrics = {
        "CPU Kullanim Yuzdesi": [r.get("cpu_avg", 0) or 0 for r in trend_data],
        "Bellek Kullanim Yuzdesi": [r.get("memory_avg", 0) or 0 for r in trend_data],
        "Disk Kullanim Yuzdesi": [r.get("disk_avg", 0) or 0 for r in trend_data],
    }

    for metric_name, values in metrics.items():
        fig, ax = plt.subplots(figsize=(12, 4), dpi=100)
        ax.plot(dates, values, marker='o', linewidth=2, markersize=6, color='#1890ff')
        ax.fill_between(dates, values, alpha=0.1, color='#1890ff')
        ax.set_xlabel("Tarih", fontsize=12)
        ax.set_ylabel("%", fontsize=12)
        ax.set_title(f"Haftalik {metric_name}", fontsize=14, fontweight="bold")
        ax.grid(True, linestyle="--", alpha=0.5)
        ax.set_facecolor("#f8f9fa")
        fig.autofmt_xdate()
        plt.tight_layout()

        buf = BytesIO()
        fig.savefig(buf, format="png", bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        charts[metric_name] = f"data:image/png;base64,{base64.b64encode(buf.getvalue()).decode('utf-8')}"

    return charts


def generate_excel(host_summary, placement_summary, rightsizing, trend_data):
    """Excel rapor dosyasi olusturur — Dashboard + ham veri sheet'leri."""
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    title_font = Font(bold=True, size=14, color="1E293B")
    subtitle_font = Font(bold=True, size=11, color="475569")
    kpi_font = Font(bold=True, size=22, color="1E293B")
    kpi_label_font = Font(size=9, color="64748B")
    warn_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
    warn_font = Font(bold=True, color="DC2626", size=10)
    data_font = Font(size=10, color="334155")

    def write_data_header(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(col)].width = max(len(h) + 4, 14)

    # =========================================================
    # Sheet 1: DASHBOARD
    # =========================================================
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_properties.tabColor = "3B82F6"

    # Baslik
    ws_dash.merge_cells("A1:H1")
    c = ws_dash.cell(row=1, column=1, value="Safir Bulut - Haftalik Kapasite Raporu")
    c.font = title_font

    ws_dash.merge_cells("A2:H2")
    ws_dash.cell(row=2, column=1, value="Otomatik olusturulmus rapor ozeti").font = Font(size=10, color="94A3B8")

    # --- KPI Kartlari (Satir 4) ---
    ws_dash.cell(row=4, column=1, value="SANAL MAKINE OZETI").font = subtitle_font

    kpi_data = [
        ("Toplam VM", rightsizing["total"], "1E293B"),
        ("Atil", len(rightsizing["idle"]), "D97706"),
        ("Fazla Kaynak", len(rightsizing["over_provisioned"]), "DC2626"),
        ("Az Kaynak", len(rightsizing["under_provisioned"]), "7C3AED"),
        ("Uygun", rightsizing["right_sized_count"], "16A34A"),
    ]
    for col, (label, val, color) in enumerate(kpi_data, 1):
        cell = ws_dash.cell(row=5, column=col, value=val)
        cell.font = Font(bold=True, size=24, color=color)
        cell.alignment = __import__('openpyxl.styles', fromlist=['Alignment']).Alignment(horizontal='center')
        lbl = ws_dash.cell(row=6, column=col, value=label)
        lbl.font = kpi_label_font
        lbl.alignment = __import__('openpyxl.styles', fromlist=['Alignment']).Alignment(horizontal='center')
        ws_dash.column_dimensions[get_column_letter(col)].width = 16

    # --- Host Kullanim Tablosu (Satir 8) ---
    ws_dash.cell(row=8, column=1, value="FIZIKSEL HOST KULLANIMI (%)").font = subtitle_font

    dash_headers = ["Host", "CPU Ort", "CPU Maks", "Memory Ort", "Memory Maks", "Disk Ort", "Disk Maks"]
    write_data_header(ws_dash, dash_headers, row=9)

    for i, h in enumerate(host_summary, 10):
        ws_dash.cell(row=i, column=1, value=h.get("nodename")).font = Font(bold=True, size=10, color="1E293B")
        cpu_avg = h.get("cpu_avg") or 0
        mem_avg = h.get("memory_avg") or 0
        disk_avg = h.get("disk_avg") or 0

        for col, val in [(2, h.get("cpu_avg")), (3, h.get("cpu_max")),
                          (4, h.get("memory_avg")), (5, h.get("memory_max")),
                          (6, h.get("disk_avg")), (7, h.get("disk_max"))]:
            cell = ws_dash.cell(row=i, column=col, value=val)
            # Hangi kaynak grubunda yuksek kullanim var?
            is_high = (col in (2, 3) and cpu_avg > 80) or (col in (4, 5) and mem_avg > 80) or (col in (6, 7) and disk_avg > 80)
            if is_high:
                cell.font = warn_font
                cell.fill = warn_fill
            else:
                cell.font = data_font

    # CPU / Memory / Disk bar charts
    chart_row = 10 + len(host_summary) + 2
    if host_summary:
        from openpyxl.chart.series import DataPoint

        chart_configs = [
            {"title": "Host CPU Kullanimi", "col": 2, "color": "3B82F6", "anchor_col": "A"},
            {"title": "Host Memory Kullanimi", "col": 4, "color": "8B5CF6", "anchor_col": "D"},
            {"title": "Host Disk Kullanimi", "col": 6, "color": "F59E0B", "anchor_col": "G"},
        ]

        # Her host icin farkli renk paleti
        host_colors = ["2563EB", "7C3AED", "DC2626", "059669", "D97706", "0891B2", "BE185D", "4338CA"]

        for cfg in chart_configs:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = cfg["title"]
            chart.y_axis.title = "%"
            chart.y_axis.scaling.max = 100
            chart.width = 16
            chart.height = 12
            chart.legend = None  # Legend kapat — host isimleri x ekseninde

            data_ref = Reference(ws_dash, min_col=cfg["col"], min_row=9, max_row=9 + len(host_summary), max_col=cfg["col"])
            cats = Reference(ws_dash, min_col=1, min_row=10, max_row=9 + len(host_summary))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)

            # Her bar'a farkli renk ata
            series = chart.series[0]
            for idx in range(len(host_summary)):
                pt = DataPoint(idx=idx)
                color = host_colors[idx % len(host_colors)]
                pt.graphicalProperties.solidFill = color
                series.data_points.append(pt)

            ws_dash.add_chart(chart, f"{cfg['anchor_col']}{chart_row}")

    # --- Placement Ozet (Chart'larin altinda) ---
    if placement_summary:
        place_start = chart_row + 16  # Chart yuksekligi ~14 satir + bosluk
        ws_dash.cell(row=place_start, column=5, value="PLACEMENT TAHSISI").font = subtitle_font
        place_headers = ["Host", "vCPU", "Memory GB", "Disk GB"]
        for col, h in enumerate(place_headers, 5):
            cell = ws_dash.cell(row=place_start + 1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, p in enumerate(placement_summary):
            r = place_start + 2 + i
            ws_dash.cell(row=r, column=5, value=p.get("nodename")).font = Font(bold=True, size=10)
            vcpu_used = p.get("vcpu_used") or 0
            vcpu_total = p.get("vcpu_total") or 1
            ws_dash.cell(row=r, column=6, value=f"{int(vcpu_used)} / {int(vcpu_total)}").font = data_font
            mem_used = p.get("memory_used_gb") or 0
            mem_total = p.get("memory_total_gb") or 1
            ws_dash.cell(row=r, column=7, value=f"{mem_used} / {mem_total}").font = data_font
            disk_used = p.get("disk_used_gb") or 0
            disk_total = p.get("disk_total_gb") or 1
            ws_dash.cell(row=r, column=8, value=f"{int(disk_used)} / {int(disk_total)}").font = data_font

    # =========================================================
    # Sheet 2: Host Metrikleri (ham veri)
    # =========================================================
    ws1 = wb.create_sheet("Host Metrikleri")
    ws1.sheet_properties.tabColor = "10B981"
    write_data_header(ws1, ["Host", "CPU Avg%", "CPU Max%", "CPU Min%",
                             "Memory Avg%", "Memory Max%", "Memory Min%",
                             "Disk Avg%", "Disk Max%", "Disk Min%"])
    for i, h in enumerate(host_summary, 2):
        ws1.cell(row=i, column=1, value=h.get("nodename"))
        ws1.cell(row=i, column=2, value=h.get("cpu_avg"))
        ws1.cell(row=i, column=3, value=h.get("cpu_max"))
        ws1.cell(row=i, column=4, value=h.get("cpu_min"))
        ws1.cell(row=i, column=5, value=h.get("memory_avg"))
        ws1.cell(row=i, column=6, value=h.get("memory_max"))
        ws1.cell(row=i, column=7, value=h.get("memory_min"))
        ws1.cell(row=i, column=8, value=h.get("disk_avg"))
        ws1.cell(row=i, column=9, value=h.get("disk_max"))
        ws1.cell(row=i, column=10, value=h.get("disk_min"))

    # =========================================================
    # Sheet 3: Placement (ham veri)
    # =========================================================
    ws2 = wb.create_sheet("Placement")
    ws2.sheet_properties.tabColor = "8B5CF6"
    write_data_header(ws2, ["Host", "vCPU Used", "vCPU Total", "Memory Used GB",
                             "Memory Total GB", "Disk Used GB", "Disk Total GB"])
    for i, p in enumerate(placement_summary, 2):
        ws2.cell(row=i, column=1, value=p.get("nodename"))
        ws2.cell(row=i, column=2, value=p.get("vcpu_used"))
        ws2.cell(row=i, column=3, value=p.get("vcpu_total"))
        ws2.cell(row=i, column=4, value=p.get("memory_used_gb"))
        ws2.cell(row=i, column=5, value=p.get("memory_total_gb"))
        ws2.cell(row=i, column=6, value=p.get("disk_used_gb"))
        ws2.cell(row=i, column=7, value=p.get("disk_total_gb"))

    # =========================================================
    # Sheet 4: VM Rightsizing (ham veri)
    # =========================================================
    ws3 = wb.create_sheet("VM Rightsizing")
    ws3.sheet_properties.tabColor = "F59E0B"
    write_data_header(ws3, ["Kategori", "VM Adi", "Instance ID", "Project ID",
                             "Avg CPU%", "Max CPU%", "Avg Memory%", "Max Memory%",
                             "vCPU", "Memory GB"])

    cat_fills = {
        "Atil": PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid"),
        "Fazla Kaynak": PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid"),
        "Az Kaynak": PatternFill(start_color="FAF5FF", end_color="FAF5FF", fill_type="solid"),
    }

    row = 2
    for category, vms in [("Atil", rightsizing["idle"]),
                           ("Fazla Kaynak", rightsizing["over_provisioned"]),
                           ("Az Kaynak", rightsizing["under_provisioned"])]:
        fill = cat_fills.get(category)
        for vm in vms:
            for col, val in [(1, category), (2, vm.get("instance_name")),
                              (3, vm.get("instance_id")), (4, vm.get("project_id")),
                              (5, vm.get("avg_cpu")), (6, vm.get("max_cpu")),
                              (7, vm.get("avg_memory")), (8, vm.get("max_memory")),
                              (9, vm.get("allocated_vcpu")), (10, vm.get("allocated_memory_gb"))]:
                cell = ws3.cell(row=row, column=col, value=val)
                if fill:
                    cell.fill = fill
            row += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
